"""Parse AI_TRADE_HIGH_RELEVANCE_PRODUCTS.md and export each category table
to its own sheet in an Excel workbook."""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "AI_TRADE_HIGH_RELEVANCE_PRODUCTS.md"
DEST = ROOT / "AI_TRADE_HIGH_RELEVANCE_PRODUCTS.xlsx"

# Categories that have a table, in document order.
CATEGORY_HEADINGS = [
    "Compute Hardware",
    "Electrical Power",
    "Networking & Telecom",
    "Cooling & HVAC",
    "Building & Structure",
    "Specialty Materials",
    "Fire Safety & Security",
]

# Excel sheet names cannot exceed 31 chars and cannot contain : \ / ? * [ ]
INVALID_SHEET_CHARS = re.compile(r"[:\\/\?\*\[\]]")


def sanitize_sheet_name(name: str) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("-", name).strip()
    return cleaned[:31]


def split_table_row(line: str) -> list[str]:
    # Strip the leading/trailing pipes, then split on remaining pipes.
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [cell.strip() for cell in stripped.split("|")]


def is_separator_row(cells: list[str]) -> bool:
    return all(re.fullmatch(r":?-+:?", cell) for cell in cells if cell)


def strip_inline_code(value: str) -> str:
    # Cells wrap HS codes in backticks; remove them so Excel treats the value as text.
    return value.replace("`", "").strip()


def parse_categories(markdown_text: str) -> dict[str, tuple[list[str], list[list[str]]]]:
    """Return a mapping of category -> (headers, rows)."""
    lines = markdown_text.splitlines()
    results: dict[str, tuple[list[str], list[list[str]]]] = {}

    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("## "):
            heading = line[3:].strip()
            if heading in CATEGORY_HEADINGS:
                # Walk forward until we find the table header row.
                j = i + 1
                while j < len(lines) and not lines[j].lstrip().startswith("|"):
                    j += 1
                if j >= len(lines):
                    i += 1
                    continue

                header_cells = split_table_row(lines[j])
                j += 1
                # Skip the separator row.
                if j < len(lines):
                    sep_cells = split_table_row(lines[j])
                    if is_separator_row(sep_cells):
                        j += 1

                rows: list[list[str]] = []
                while j < len(lines):
                    candidate = lines[j]
                    if not candidate.lstrip().startswith("|"):
                        break
                    cells = [strip_inline_code(c) for c in split_table_row(candidate)]
                    rows.append(cells)
                    j += 1

                results[heading] = (header_cells, rows)
                i = j
                continue
        i += 1

    return results


def write_workbook(parsed: dict[str, tuple[list[str], list[list[str]]]], dest: Path) -> None:
    wb = Workbook()
    # Remove the default sheet; we'll add per-category sheets in document order.
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for category in CATEGORY_HEADINGS:
        if category not in parsed:
            continue
        headers, rows = parsed[category]
        sheet = wb.create_sheet(title=sanitize_sheet_name(category))

        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row in rows:
            # Pad/truncate to header length so columns align.
            normalized = (row + [""] * len(headers))[: len(headers)]
            sheet.append(normalized)

        for row_cells in sheet.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = body_alignment

        # Reasonable column widths: HS code narrow, prose columns wide.
        widths = [16, 45, 70, 60]
        for idx in range(len(headers)):
            width = widths[idx] if idx < len(widths) else 30
            sheet.column_dimensions[get_column_letter(idx + 1)].width = width

        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 28

    wb.save(dest)


def main() -> None:
    text = SRC.read_text(encoding="utf-8")
    parsed = parse_categories(text)

    print(f"Parsed {len(parsed)} categories from {SRC.name}:")
    for category in CATEGORY_HEADINGS:
        if category in parsed:
            _, rows = parsed[category]
            print(f"  - {category}: {len(rows)} rows")
        else:
            print(f"  - {category}: NOT FOUND")

    write_workbook(parsed, DEST)
    print(f"\nWrote {DEST}")


if __name__ == "__main__":
    main()
